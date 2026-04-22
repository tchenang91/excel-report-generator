"""Excel Report Generator - Automated formatted Excel reports."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

HEADER_FILL = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=10)
GOLD = PatternFill(start_color='C9A84C', end_color='C9A84C', fill_type='solid')
BORDER = Border(
    left=Side(style='thin', color='D4D4D4'),
    right=Side(style='thin', color='D4D4D4'),
    top=Side(style='thin', color='D4D4D4'),
    bottom=Side(style='thin', color='D4D4D4')
)

def generate_demo_data():
    return pd.DataFrame({
        'Monat': ['Jan','Feb','Mär','Apr','Mai','Jun','Jul','Aug','Sep','Okt','Nov','Dez'],
        'Umsatz': [125000,132000,128000,145000,152000,148000,160000,155000,168000,172000,180000,195000],
        'Kosten': [98000,101000,97000,108000,112000,109000,118000,114000,123000,126000,131000,140000],
        'Gewinn': [27000,31000,31000,37000,40000,39000,42000,41000,45000,46000,49000,55000]
    })

def create_report(df, output='report.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Monatsbericht'
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Monatsbericht — {datetime.now().strftime("%B %Y")}'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1B4F72')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Data table
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
            if r_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            else:
                cell.font = DATA_FONT
                if c_idx >= 2:
                    cell.number_format = '#,##0 €'
    
    # Auto-width
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    
    # Chart
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Umsatz vs. Kosten'
    chart.y_axis.title = 'EUR'
    data = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=15)
    cats = Reference(ws, min_col=1, min_row=4, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, 'F3')
    
    wb.save(output)
    print(f'✓ Report saved: {output}')

if __name__ == '__main__':
    df = generate_demo_data()
    create_report(df)
